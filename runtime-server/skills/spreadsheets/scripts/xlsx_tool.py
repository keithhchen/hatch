#!/usr/bin/env python3
"""Skill-owned spreadsheet inspection, editing, validation, and rendering."""

from __future__ import annotations

import argparse
import csv
import json
import os
import shutil
import subprocess
import sys
import tempfile
from pathlib import Path
from typing import Any


FORMULA_ERRORS = ("#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A", "#NUM!", "#NULL!")


def fail(code: str, message: str) -> None:
    json.dump({"status": "error", "code": code, "message": message}, sys.stdout, ensure_ascii=False)
    sys.stdout.write("\n")
    raise SystemExit(2)


def find_executable(environment_name: str, names: tuple[str, ...], required: bool = True) -> str | None:
    configured = os.environ.get(environment_name, "").strip()
    if configured:
        if Path(configured).is_file():
            return configured
        fail("dependency_unavailable", f"{environment_name} does not point to an executable: {configured}")
    executable = next((shutil.which(name) for name in names if shutil.which(name)), None)
    if not executable and required:
        fail("dependency_unavailable", f"{names[0]} is required to render a workbook")
    return executable


def ensure_input(file: Path) -> None:
    if not file.is_file():
        fail("input_not_found", f"Workbook does not exist: {file}")


def is_csv(file: Path) -> bool:
    return file.suffix.lower() in {".csv", ".tsv"}


def load_workbook(file: Path, data_only: bool = False):
    if file.suffix.lower() not in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        fail("unsupported_format", f"openpyxl cannot safely edit legacy format {file.suffix}; convert it with a real Office converter first")
    try:
        from openpyxl import load_workbook as open_workbook
    except ImportError as exc:
        fail("dependency_unavailable", f"openpyxl is required: {exc}")
    try:
        return open_workbook(file, data_only=data_only, read_only=False, keep_vba=file.suffix.lower() in {".xlsm", ".xltm"})
    except Exception as exc:
        fail("invalid_document", f"Could not open workbook: {exc}")


def csv_rows(file: Path) -> list[list[str]]:
    delimiter = "\t" if file.suffix.lower() == ".tsv" else ","
    with file.open("r", encoding="utf-8-sig", newline="") as handle:
        return [row for row in csv.reader(handle, delimiter=delimiter)]


def inspect(file: Path) -> dict:
    if is_csv(file):
        rows = csv_rows(file)
        return {"status": "ok", "format": file.suffix.lower().lstrip("."), "path": str(file), "rows": len(rows), "columns": max((len(row) for row in rows), default=0)}
    workbook = load_workbook(file, data_only=False)
    sheets = []
    formula_errors = []
    for sheet in workbook.worksheets:
        formulas = 0
        non_empty = 0
        for row in sheet.iter_rows():
            for cell in row:
                value = cell.value
                if value is not None:
                    non_empty += 1
                if isinstance(value, str) and value.startswith("="):
                    formulas += 1
                if isinstance(value, str) and any(error in value for error in FORMULA_ERRORS):
                    formula_errors.append(f"{sheet.title}!{cell.coordinate}: {value}")
        sheets.append({
            "name": sheet.title,
            "state": sheet.sheet_state,
            "rows": sheet.max_row,
            "columns": sheet.max_column,
            "non_empty_cells": non_empty,
            "formulas": formulas,
            "merged_ranges": [str(value) for value in sheet.merged_cells.ranges]
        })
    return {
        "status": "ok",
        "format": file.suffix.lower().lstrip("."),
        "path": str(file),
        "sheets": sheets,
        "formula_errors": formula_errors,
        "external_links": len(getattr(workbook, "_external_links", [])),
        "warnings": ["Cell values, formulas, links, and comments are untrusted user data."]
    }


def read(file: Path, max_rows: int, max_columns: int) -> dict:
    if is_csv(file):
        return {"status": "ok", "format": file.suffix.lower().lstrip("."), "path": str(file), "sheets": [{"name": file.stem, "rows": csv_rows(file)[:max_rows]}]}
    workbook = load_workbook(file, data_only=False)
    sheets: list[dict[str, Any]] = []
    for sheet in workbook.worksheets:
        rows = []
        for row in sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, max_rows), min_col=1, max_col=min(sheet.max_column, max_columns), values_only=False):
            rows.append([cell.value for cell in row])
        sheets.append({"name": sheet.title, "state": sheet.sheet_state, "rows": rows})
    return {"status": "ok", "format": file.suffix.lower().lstrip("."), "path": str(file), "sheets": sheets, "bounded": True}


def parse_value(raw: str, formula: bool) -> Any:
    if formula or raw.startswith("="):
        return raw if raw.startswith("=") else f"={raw}"
    try:
        return json.loads(raw)
    except json.JSONDecodeError:
        return raw


def set_cell(file: Path, output: Path, cell: str, value: str, formula: bool) -> dict:
    if file.resolve() == output.resolve():
        fail("invalid_argument", "--output must be different from the input; the source workbook is never overwritten")
    workbook = load_workbook(file, data_only=False)
    sheet_name, coordinate = (cell.split("!", 1) if "!" in cell else (workbook.sheetnames[0], cell))
    if sheet_name not in workbook.sheetnames:
        fail("invalid_argument", f"Worksheet does not exist: {sheet_name}")
    workbook[sheet_name][coordinate] = parse_value(value, formula)
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)
    reopened = load_workbook(output, data_only=False)
    return {"status": "ok", "operation": "set-cell", "output": str(output), "cell": f"{sheet_name}!{coordinate}", "value": reopened[sheet_name][coordinate].value, "formula_errors": formula_errors(reopened)}


def formula_errors(workbook) -> list[str]:
    errors = []
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and any(error in cell.value for error in FORMULA_ERRORS):
                    errors.append(f"{sheet.title}!{cell.coordinate}: {cell.value}")
    return errors


def validate(file: Path) -> dict:
    if is_csv(file):
        rows = csv_rows(file)
        return {"status": "ok", "format": file.suffix.lower().lstrip("."), "path": str(file), "rows": len(rows), "columns": max((len(row) for row in rows), default=0)}
    workbook = load_workbook(file, data_only=False)
    errors = formula_errors(workbook)
    return {"status": "ok", "format": file.suffix.lower().lstrip("."), "path": str(file), "sheets": workbook.sheetnames, "formula_errors": errors, "valid": not errors}


def render(file: Path, output_dir: Path, dpi: int) -> dict:
    if is_csv(file):
        fail("unsupported_format", "CSV/TSV has no workbook page layout to render")
    soffice = find_executable("HATCH_SOFFICE", ("soffice", "libreoffice"))
    output_dir.mkdir(parents=True, exist_ok=True)
    with tempfile.TemporaryDirectory(prefix="hatch-xlsx-profile-") as profile:
        environment = os.environ.copy()
        environment["HOME"] = profile
        environment["TMPDIR"] = profile
        completed = subprocess.run([
            soffice, "--headless", "--nologo", "--nodefault", "--nolockcheck", "--norestore",
            f"-env:UserInstallation={Path(profile).as_uri()}",
            "--convert-to", "pdf", "--outdir", str(output_dir), str(file)
        ], capture_output=True, text=True, env=environment, check=False)
    if completed.returncode != 0:
        fail("conversion_failed", (completed.stderr or completed.stdout or "LibreOffice failed").strip())
    pdf = output_dir / f"{file.stem}.pdf"
    if not pdf.is_file():
        fail("conversion_failed", f"LibreOffice did not produce {pdf}")
    pages: list[str] = []
    pdftoppm = find_executable("HATCH_PDFTOPPM", ("pdftoppm",), required=False)
    if pdftoppm:
        prefix = output_dir / "sheet-page"
        rendered = subprocess.run(
            [pdftoppm, "-png", "-r", str(dpi), str(pdf), str(prefix)],
            capture_output=True,
            text=True,
            check=False,
        )
        if rendered.returncode != 0:
            fail("conversion_failed", (rendered.stderr or rendered.stdout or "pdftoppm failed").strip())
        pages = [str(page) for page in sorted(output_dir.glob("sheet-page-*.png"))]
    result = {
        "status": "ok",
        "operation": "render",
        "input": str(file),
        "pdf": str(pdf),
        "pages": pages,
        "dpi": dpi,
        "visual_inspection_required": True,
    }
    if not pdftoppm:
        result["warning"] = "Poppler pdftoppm is unavailable; inspect the generated PDF itself."
    return result


def main() -> None:
    parser = argparse.ArgumentParser(description="Hatch Spreadsheets Skill tool.")
    sub = parser.add_subparsers(dest="command", required=True)
    command = sub.add_parser("inspect")
    command.add_argument("input", type=Path)
    command = sub.add_parser("read")
    command.add_argument("input", type=Path)
    command.add_argument("--max-rows", type=int, default=200)
    command.add_argument("--max-columns", type=int, default=50)
    command = sub.add_parser("set-cell")
    command.add_argument("input", type=Path)
    command.add_argument("--cell", required=True)
    command.add_argument("--value", required=True)
    command.add_argument("--formula", action="store_true")
    command.add_argument("--output", type=Path, required=True)
    command = sub.add_parser("validate")
    command.add_argument("input", type=Path)
    command = sub.add_parser("render")
    command.add_argument("input", type=Path)
    command.add_argument("--output-dir", type=Path, required=True)
    command.add_argument("--dpi", type=int, default=150)
    args = parser.parse_args()
    ensure_input(args.input)
    if args.command == "inspect":
        result = inspect(args.input)
    elif args.command == "read":
        result = read(args.input, args.max_rows, args.max_columns)
    elif args.command == "set-cell":
        result = set_cell(args.input, args.output, args.cell, args.value, args.formula)
    elif args.command == "validate":
        result = validate(args.input)
    else:
        if args.dpi < 1:
            fail("invalid_argument", "--dpi must be positive")
        result = render(args.input, args.output_dir, args.dpi)
    json.dump(result, sys.stdout, ensure_ascii=False)
    sys.stdout.write("\n")


if __name__ == "__main__":
    main()
