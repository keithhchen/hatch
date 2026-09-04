#!/usr/bin/env python3
"""Recalculate an OOXML workbook through LibreOffice and verify cached values."""

from __future__ import annotations

import argparse
import json
import os
import shutil
import subprocess
import sys
import tempfile
from pathlib import Path
from typing import Any

sys.path.insert(0, str(Path(__file__).resolve().parents[2]))
from _shared.libreoffice import run_libreoffice


FORMULA_ERRORS = ("#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A", "#NUM!", "#NULL!")
SUPPORTED = {".xlsx", ".xlsm", ".xltx", ".xltm"}


def fail(code: str, message: str) -> None:
    json.dump({"status": "error", "code": code, "message": message}, sys.stdout, ensure_ascii=False)
    sys.stdout.write("\n")
    raise SystemExit(2)


def ensure_input(file: Path) -> None:
    if not file.is_file():
        fail("input_not_found", f"Workbook does not exist: {file}")
    if file.suffix.lower() not in SUPPORTED:
        fail("unsupported_format", f"LibreOffice recalculation supports OOXML workbooks, not {file.suffix or 'this format'}")


def find_soffice() -> str:
    executable = os.environ.get("HATCH_SOFFICE", "").strip()
    if executable:
        if Path(executable).is_file():
            return executable
        fail("dependency_unavailable", f"HATCH_SOFFICE does not point to an executable: {executable}")
    executable = shutil.which("soffice") or shutil.which("libreoffice")
    if not executable:
        fail("dependency_unavailable", "LibreOffice (soffice) is required to recalculate a workbook")
    return executable


def load_workbook(file: Path, data_only: bool):
    try:
        from openpyxl import load_workbook as open_workbook
    except ImportError as exc:
        fail("dependency_unavailable", f"openpyxl is required: {exc}")
    try:
        return open_workbook(
            file,
            data_only=data_only,
            read_only=False,
            keep_vba=file.suffix.lower() in {".xlsm", ".xltm"},
        )
    except Exception as exc:
        fail("invalid_document", f"Could not open workbook: {exc}")


def formula_cells(workbook) -> list[str]:
    cells: list[str] = []
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    cells.append(f"{sheet.title}!{cell.coordinate}")
    return cells


def value_errors(workbook) -> list[str]:
    errors: list[str] = []
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, str) and any(error in value for error in FORMULA_ERRORS):
                    errors.append(f"{sheet.title}!{cell.coordinate}: {value}")
    return errors


def recalculate(source: Path, output: Path) -> dict[str, Any]:
    if source.resolve() == output.resolve():
        fail("invalid_argument", "--output must be different from the input; the source workbook is never overwritten")
    if output.suffix.lower() != ".xlsx":
        fail("unsupported_format", "Recalculation output must be .xlsx so LibreOffice preserves the OOXML workbook contract")
    soffice = find_soffice()
    before = load_workbook(source, data_only=False)
    formulas = formula_cells(before)
    before_errors = value_errors(before)

    with tempfile.TemporaryDirectory(prefix="hatch-xlsx-recalc-") as workspace:
        workspace_path = Path(workspace)
        profile = workspace_path / "profile"
        input_dir = workspace_path / "input"
        output_dir = workspace_path / "output"
        input_dir.mkdir()
        output_dir.mkdir()
        # Work on a copy. LibreOffice may update links, calculation metadata,
        # or other package state while opening a workbook.
        staged_input = input_dir / f"source{source.suffix.lower()}"
        shutil.copy2(source, staged_input)
        environment = os.environ.copy()
        environment["HOME"] = str(profile)
        environment["TMPDIR"] = str(workspace_path / "tmp")
        (workspace_path / "tmp").mkdir()
        command = [
            soffice,
            "--headless",
            "--invisible",
            "--nologo",
            "--nodefault",
            "--nolockcheck",
            "--norestore",
            "--safe-mode",
            f"-env:UserInstallation={profile.as_uri()}",
            "--convert-to",
            "xlsx:Calc MS Excel 2007 XML",
            "--outdir",
            str(output_dir),
            str(staged_input),
        ]
        completed = run_libreoffice(command, profile=profile, environment=environment)
        if completed.returncode != 0:
            fail("conversion_failed", (completed.stderr or completed.stdout or "LibreOffice recalculation failed").strip())
        recalculated = output_dir / "source.xlsx"
        if not recalculated.is_file():
            fail("conversion_failed", f"LibreOffice did not produce {recalculated}")
        output.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(recalculated, output)

    after_formula = load_workbook(output, data_only=False)
    after_values = load_workbook(output, data_only=True)
    formula_errors = value_errors(after_formula)
    cached_errors = value_errors(after_values)
    cached_values_available = True
    for reference in formulas:
        sheet_name, coordinate = reference.split("!", 1)
        if after_values[sheet_name][coordinate].value is None:
            cached_values_available = False
            break
    return {
        "status": "ok",
        "operation": "recalculate",
        "input": str(source),
        "output": str(output),
        "renderer": "LibreOffice",
        "formula_count": len(formulas),
        "formula_errors_before": before_errors,
        "formula_errors": formula_errors,
        "cached_formula_errors": cached_errors,
        "cached_values_available": cached_values_available,
        "valid": not formula_errors and not cached_errors,
        "stdout": completed.stdout.strip(),
        "warning": "Macros were not executed. External links remain untrusted user data.",
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Recalculate an Excel workbook through LibreOffice.")
    parser.add_argument("input", type=Path)
    parser.add_argument("--output", type=Path, required=True)
    args = parser.parse_args()
    ensure_input(args.input)
    result = recalculate(args.input, args.output)
    json.dump(result, sys.stdout, ensure_ascii=False)
    sys.stdout.write("\n")


if __name__ == "__main__":
    main()
